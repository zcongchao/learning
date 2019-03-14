# -*- coding: utf-8 -*-
"""
Created on Wed Nov 28 09:51:42 2018

@author: Administrator
"""

import codecs
import datetime
from bs4 import BeautifulSoup

from selenium import webdriver
import selenium.webdriver.support.ui as ui
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time

from lxml import etree
from HTMLParser import HTMLParser

import openpyxl



class JdSpider(object):

    def get_goods(self,driver):#, goods
    
        select = driver.find_element_by_xpath("//a[@href='http://finance.sina.com.cn/futuremarket/spotprice.shtml#titlePos_1']")
        select.click()
        time.sleep(3)  
        
        driver.switch_to_window(driver.window_handles[1])
        page_num=0
        
        """
        ele_toname = driver.find_element_by_id('key')
        
        ele_search = driver.find_element_by_class_name('button')
        
        
     
        ele_toname.clear()
        ele_toname.send_keys(goods)
        ele_toname.click()
     
        
        
        ele_search.click()
        page_num=0
   
        while True:
            try:
                WebDriverWait(driver, 10).until(
                    EC.title_contains(unicode(goods))
                )
            except Exception,e:
                print e
                break
            time.sleep(5)
     
            js = "window.scrollTo(0, document.body.scrollHeight);"
            driver.execute_script(js)
            time.sleep(5)
            """
      
        btc = driver.find_element_by_id("tab_switch_3")
        btc.click()
        time.sleep(3)
        
        
        xf = driver.find_element_by_xpath("//iframe[@src='http://www.100ppi.com/price/table2.html']")
        driver.switch_to.frame(xf)
        
        htm_const = driver.page_source
        soup = BeautifulSoup(htm_const,'html.parser', from_encoding='utf-8')
            
            
        infos = str(soup.find_all(class_="ftab"))
        
        tree = etree.HTML(infos)
        
        data = [3200]
        nums = [33,32,2,37,35,11,36,26,40,34,38,30,39,25,29,31,19,21]
        for num in nums:
            content1 = tree.xpath("//tr/td[last()-6]/text()")[num].decode('unicode_escape')
            #去掉空格字符和小数位数
            content1="".join(content1.split())
            content1=float(content1)
            content1 = round(content1)
            data.append(content1)
            
        workbook=openpyxl.load_workbook("D:\kuapingzhongcanshu.xlsx")

        #sheet_ranges =workbook['Sheet1']
        #print sheet_ranges['A1'].value 
        ws = workbook['Sheet1'] 
    
        num = ['3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21']
    
    
        for num1, value1 in zip(num, data):
    
            col = 'D' + num1
            ws[col] = value1
    #ws["A1"] = '100000000000000000'   
     
    
        workbook.save(filename="D:/new2.xlsx")
            #print data
            
            #print content1
            
        
        
        """
        f = codecs.open(u'sina.html', 'a', 'utf-8')
        for info in infos:
            f.write(str(page_num)+'--'*50)
            content = info.get_text().replace(" ","").replace("\t","").strip()
            for line in [ln for ln in content.splitlines() if ln.strip()]:
                f.write(line)
                f.write('\r\n')
        f.close()
        """
        
        #print infos
      
        
        
            

        """
        htm_const = driver.page_source
        soup = BeautifulSoup(htm_const,'html.parser', from_encoding='utf-8')
        
        
        infos = soup.find_all(class_="gl-item")
        
        
        f = codecs.open(unicode(goods)+ u'.html', 'a', 'utf-8')
        for info in infos:
            f.write(str(page_num)+'--'*50)
            content = info.get_text().replace(" ","").replace("\t","").strip()
            for line in [ln for ln in content.splitlines() if ln.strip()]:
                f.write(line)
                f.write('\r\n')
        f.close()
        try:
            next_page = WebDriverWait(driver, 10).until(
            EC.visibility_of(driver.find_element_by_css_selector(".item.next"))
            )
            next_page.click()
            page_num+=1
            time.sleep(10)
        except Exception,e:
            print e
            break
        """

    def crawl(self,root_url):#, goods
         
         driver = webdriver.Firefox(executable_path='D:\Anaconda2\Scripts\geckodriver.exe')
         driver.set_page_load_timeout(50)
         driver.get(root_url)
         driver.maximize_window() # 将浏览器最大化显示
         driver.implicitly_wait(10) # 控制间隔时间，等待浏览器反映
         self.get_goods(driver)


if __name__=='__main__':
    spider = JdSpider()
    spider.crawl('https://finance.sina.com.cn/futuremarket/')#,u"thinkpad"
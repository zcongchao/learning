# -*- coding: utf-8 -*-

import sys

if sys.getdefaultencoding() != 'utf-8':
    reload(sys)
    sys.setdefaultencoding('utf-8')


import openpyxl
 
workbook=openpyxl.load_workbook("D:\kuapingzhongcanshu.xlsx")

sheet_ranges =workbook['Sheet1']
#print sheet_ranges['A1'].value 
ws = workbook['Sheet1'] 

num = ['3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21']
value = ['3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21']

for num1, value1 in zip(num, value):

    col = 'D' + num1
    ws[col] = value1
#ws["A1"] = '100000000000000000'   
 

workbook.save(filename="D:/new2.xlsx")
